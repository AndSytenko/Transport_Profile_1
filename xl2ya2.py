#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''
Convers csv files (KS MEN21 project) into a yaml variable structure
'''

__author__ = 'oleg.larintsev@nokia.com'
__version__ = '0.2.0'

import yaml
import getopt
import sys
import os
import unicodedata
import re
from os.path import basename
from jxlib import *
from openpyxl import load_workbook
dbg = 0

def redopt():
    o = {}
    try:
        opts, args = getopt.getopt(sys.argv[1:],"dha:r:i:y:",["aggrs=","rings=", "index=", "yaml="])
    except getopt.GetoptError as e:
        error (str(e))
        usage()
        sys.exit(2)

    if not opts:
        usage()
        quit()

    for opt, arg in opts:
        if opt == '-h':
            usage()
            sys.exit()
        elif opt in ('-a', '--aggrs'):
            o['aggrs'] = arg
        elif opt in ('-r', '--rings'):
            o['rings'] = arg
        elif opt in ('-i', '--index'):
            o['index'] = arg
        elif opt in ('-y', '--yaml'):
            o['yaml'] = arg
        elif opt in ('-d', '--debug'):
            dbg = 1

    if not 'aggrs' in o:
        error('The \"-a\" switch is mandatory!')
        sys.exit(2)
    if not 'rings' in o:
        error('The \"-r\" switch is mandatory!')
        sys.exit(2)
    #if not 'index' in o:
    #    error('The \"-i\" switch is mandatory!')
    #    sys.exit(2)
    if 'index' in o:
        try:
            int(o['index'])
        except ValueError:
            error("Index must be number!")
            sys.exit(2)
    return o

def usage ():
    info ("Usage: %s -a <aggregtion routers file> -r <rings file> [-i <ring index>] [-y <yaml_file>] [-h] [-d]" % sys.argv[0])
    sys.exit(2)

def get_aggr_nodes_data (_wb):
    #WS structure:
    AGGR_ID        =  0 # 0:  (#) Num         (int)
    AGGR_SYSTEM_IP =  1 # 1:  System Loopback (str)
    AGGR_NAME      =  2 # 2:  Name            (str)
    AGGR_NSID      =  3 # 3:  Node SID        (int)
    AGGR_SBFD_ID   =  4 # 4:  S-BFD Discriminator (int)
    AGGR_CHASSIS   =  5 # 5:  Chassis Type    (str)
    AGGR_SROS      =  6 # 6:  Software Version(str)
    AGGR_SDP_ID    =  7 # 7:  SDP ID          (int)

    ###    0      1               2        3      4         5            6                 7
    ### ['1', '172.16.252.1', 'sr1-kie2', '1', '524288', '7750-SR12', 'TiMOS-C-19.10.R6', 1001]
    ###   id      system_ip     name      sid   bfd        chassis      SROS              sdpid

    a_header_map = [ 'id', 'ip', 'name', 'sid', 'sbfd', 'chassis', 'sros', 'sdpid' ]
    anodes = {}
    current_row = 0

    if 'SR' not in _wb.sheetnames:
        error("'SR' worksheet was not found!")
        sys.exit(2)
    ws = wb['SR']
    info( "WS 'SR': rows: {:d}; cols: {:d}".format(ws.max_row, ws.max_column))
    for row in ws.iter_rows(max_col=8, values_only=True):
        # skip the header
        if current_row == 0:
            current_row += 1
            continue
        if not (row[AGGR_ID] == None and row[AGGR_SYSTEM_IP] == None and row[AGGR_NAME] == None):
            print ("valid!")
            node = {}
            print("-> {}, {}, {}".format(row[AGGR_ID], row[AGGR_SYSTEM_IP], row[AGGR_NAME]))

            node =  (dict(zip(a_header_map,row)))
            anodes[row[AGGR_NAME]] = node
            debug ("AGGR data: {}".format(node), dbg)

            #print("Nodes struct: {}".format(dsnodes))

        current_row += 1
    return anodes

def proc_dsnodes_list2(_wb, _ring = 0):
    #WS structure:
    NODE_NUM       =  0 # 0:  (#) Num   (int)
    NODE_RING      =  1 # 1:  Ring      (int)
    NODE_SITE      =  2 # 2:  Site      (str)
    NODE_NAME      =  3 # 3:  Node      (str)
    NODE_SYSTEM_IP =  4 # 4:  System    (str)
    NODE_NSID      =  5 # 5:  N-SID     (int)
    NODE_SBFD_ID   =  6 # 6:  SBFD ID   (int)
    NODE_AGGR      =  7 # 7:  Aggr      (str)
    NODE_OSPF_AREA      =  8 # 8:  OSPF Area (int)
    NODE_SDP_ID    =  9 # 9:  SDP ID    (int)
    NODE_CHASSIS   = 10 # 10: Chassis   (str)
    NODE_SROS      = 11 # 11: SR OS     (str)
    NODE_PAIR_ID   = 12 # 12: Pair ID   (int)

    dsnodes = {}
    dspairs = {}
    current_row = 0
    #  0     1      2      3            4         5        6       7       8          9        10        11        12
    # '#'; Ring #; Site; Node;       System;     N-SID;  SBFD ID; Aggr; OSPF Area;  SDP ID;  Chassis;   SR OS;    Pair ID
    #  1;   1;   UO0880; ds1-uo0880; 172.25.5.1; 1501;   524801;  ode1;  215;       13001;  7250 IXR-E; 20.10.R3; 10844

    def fill_node_struct():
        if row[NODE_RING] not in dsnodes.keys():
            dsnodes[row[NODE_RING]] = {}
        if row[NODE_NAME] not in dsnodes[row[NODE_RING]].keys():
            dsnodes[row[NODE_RING]][row[NODE_NAME]] = {}
        if int(row[NODE_PAIR_ID]) not in dspairs.keys():
        # add dspair params only once
            dspairs[int(row[NODE_PAIR_ID])] = {}
            dspairs[int(row[NODE_PAIR_ID])]['nodes'] = []
            dspairs[int(row[NODE_PAIR_ID])]['site'] = row[NODE_SITE]
            dspairs[int(row[NODE_PAIR_ID])]['ring'] = int(row[NODE_RING])
        dspairs[int(row[NODE_PAIR_ID])]['nodes'].append(row[NODE_NAME])
        if dspairs[int(row[NODE_PAIR_ID])]['site'] != row[NODE_SITE]:
            warning("Sites for nodes {} does not match: {} vs {}!!".format(dspairs[int(row[NODE_PAIR_ID])]['nodes'],                         dspairs[int(row[NODE_PAIR_ID])]['sites'], row[NODE_SITE]))
        node['site'] = {'name': row[NODE_SITE].strip(' '), 'site_id': -1, 'ring': int(row[NODE_RING]), 'pair_id': int(row[NODE_PAIR_ID]), 'site_peer': '', 'node_idx': 0 }
        node['bfd']['discriminator'] = int(row[NODE_SBFD_ID])
        node['ospf'].append({ 'area' : int(row[NODE_OSPF_AREA]) })
        node['sdp'] = int(row[NODE_SDP_ID])
        node['interfaces'].append( {'name': 'system', 'sid': int(row[NODE_NSID]), 'ip': "{}/32".format(row[NODE_SYSTEM_IP]), 'vlan': None, 'port': 'system', 'ospf_area': int(row[NODE_OSPF_AREA])} )
        dsnodes[row[NODE_RING]][row[NODE_NAME]].update(node)
        debug("Filling DS node structure with data: \n\tRing: {}; Node: {}; IP: {}".format(row[NODE_RING], row[NODE_NAME], row[NODE_SYSTEM_IP]), dbg)

    # End of the def

    if 'Nodes' not in _wb.sheetnames:
        error("'Nodes' sheet was not found!")
        sys.exit(2)
    ws = _wb['Nodes']
    info( "WB 'Nodes': rows: {:d}; cols: {:d}".format(ws.max_row, ws.max_column))
    for row in ws.iter_rows(max_col=13, values_only=True):
        # skip the header
        if current_row == 0:
            current_row += 1
            continue
        if not (row[NODE_NUM] == None and row[NODE_RING] == None and row[NODE_SITE] == None and row[NODE_NAME] == None):
            print ("valid!")
            node = {}
            print("-> {}, {}, {}, {}".format(row[NODE_NUM], row[NODE_RING], row[NODE_SITE], row[NODE_NAME]))
            node['bfd'] = {}
            node['ospf'] = []
            node['interfaces'] = []
            node['ports'] = []
            if _ring != 0:
                print("RING switch: {}".format(_ring))
            # Process a given ring only
                if int(row[NODE_RING]) == int(_ring):
                    print("fill node struct")
                    fill_node_struct()
            else:
            # Process all rings
                fill_node_struct()
            #print("Nodes struct: {}".format(dsnodes))

        current_row += 1

    return dsnodes, dspairs

def proc_rings_list(_rings_list):
    rl = []
    with open(_rings_list, "r") as f:
        for i, l in enumerate(f):
            if i == 0:
                continue
#            print("i: {}; Ring LINE: {}".format(i, l))
            l = re.split(';', str_norm(l.rstrip().lstrip()), 13)
            rl.append (l)
    return rl

def proc_rings_list2(_wb, _ring):
    #WS structure:
    RING_ID        =  0 # 0:  Ring      (int)
    RING_SITE_ID   =  1 # 1:  Site #    (int)
    RING_NODE_NUM  =  2 # 2:  Node #    (int)
    RING_SITE      =  3 # 3:  Site      (str)
    RING_NODE      =  4 # 4:  Node      (str)
    RING_PEER      =  5 # 5:  PeerD     (str)
    RING_PORT      =  6 # 6:  Port      (str)
    RING_PEER_PORT =  7 # 7:  Peer port (str)
    RING_PORT_TYPE =  8 # 8:  Port type (str)
    RING_SFP       =  9 # 9:  SFP/XFP   (str)
    RING_INTERFACE = 10 # 10: Interface (str)
    RING_OSPF_AREA = 11 # 11: OSPF Area (str)
    RING_IP_ADDR   = 12 # 12: IP Address(str)
    RING_COMMENT   = 13 # 13: Comment   (str)

    def fill_rings_structs():
        if row[RING_ID] not in ring_rows.keys():
            ring_rows[row[RING_ID]] = []
        ring_rows[row[RING_ID]].append(row)
        if row[RING_ID] not in rings.keys():
            rings[RING_ID] = []
        rings[RING_ID].append(row[RING_NODE])

    ring_rows = {}
    current_row = 0

    if 'Rings' not in _wb.sheetnames:
        error("'Rings' sheet was not found!", dbg)
        sys.exit(2)
    ws = _wb['Rings']
    info( "WB 'Rings': rows: {:d}; cols: {:d}".format(ws.max_row, ws.max_column))
    for row in ws.iter_rows(max_col=14, values_only=True):
        # skip the header
        if current_row == 0:
            current_row += 1
            continue
        if not (row[RING_ID] == None and row[RING_SITE_ID] == None and row[RING_NODE_NUM] == None and row[RING_SITE] == None):
            print("-> {}, {}, {}, {}".format(row[RING_ID], row[RING_SITE_ID], row[RING_NODE_NUM], row[RING_SITE]))
            if _ring != 0:
            # Process a given ring only, save data into a special variable, ring specific, then process
                if int(row[RING_ID]) == int(_ring):
                    fill_rings_structs()
            else:
            # Process all rings$
                fill_rings_structs()

            # Fill DS node routers data struct for the rings\ring N from the 'ring_rows' list
# Parse ring data
##    0      1          2           3             4           5          6            7               8           9            10            11     12
##  Ring   Site #    Node #       Site        Node          Peer        Port        Peer port    Port Media       SFP       Interface       OSPF     IP (/31)
## ['25',   '1',      '1',      'KIE3',     'sr1-kie3',   'ds1-ua0747', '', '       1/1/31',    '10GBASE-LR',   'SFP+',   'ds1-ua0747_if1', '210', '172.24.0.144;']
## ['25',   '2',      '2',      'UA0747',   'ds1-ua0747', 'sr1-kie3',   '1/1/31',   '0',        '10GBASE-LR',   'SFP+',   'sr1-kie3_if1',   '210', '172.24.0.145;']

    for rid in ring_rows.keys():
        if rid not in aggrs.keys():
            info("New ring {} for aggragators".format(rid))
            aggrs[rid] = {}
        print("Ring #{}".format(rid))
        print("\t => {}".format(ring_rows[rid]))
        for row_idx, row in enumerate (ring_rows[rid]):
            if (row_idx % 2) == 0:
                _srlg = "SRLG-ACW"
            else:
                _srlg = "SRLG-CW"

            print("RING_NODE: {}, dsnodes.keys: {}".format(row[RING_NODE], dsnodes[rid].keys()))
            if row[RING_NODE] not in dsnodes[rid].keys():
                info("Node \"{}\" is not in known as distributed: aggregator?".format(row[RING_NODE]))
                # If a node is not one of DS node check the Aggr list data
                ##if r[RING_NODE] in aggrs.keys():
                if row[RING_NODE] in daggr:
                #if row[RING_NODE] in aggr_nodes:
                    anode = {}
                    anode['interfaces'] = []
                    anode['ports'] = []
                    # It is - check if it's known yet
                    #if row[RING_NODE] not in aggrs_in_a_ring:
                    if row[RING_NODE] not in aggrs[rid].keys():
                        # Not known -> create a base structure for a node
                        aggrs[rid][row[RING_NODE]] = {}
                        anode['bfd'] = {}
                        anode['ospf'] = []
                        anode['chassis'] = "{}".format(get_chassis(daggr[row[RING_NODE]]['chassis']))
                        anode['timos'] = get_timos_version(daggr[row[RING_NODE]]['sros'])
                        anode['bfd']['discriminator'] = int(daggr[row[RING_NODE]]['sbfd'])
                        anode['ospf'].append({ 'area' : int(row[RING_OSPF_AREA]) })
                        anode['interfaces'].append({ 'name': 'system', 'sid': int(daggr[row[RING_NODE]]['sid']), 'ip': "{}/32".format(daggr[row[RING_NODE]]['ip']), 'vlan': None, 'port': 'system', 'ospf_area': 0 } )
                        aggrs_in_a_ring.append(row[RING_NODE])
                        # Append a ring data to a node (aggr) data struct
                        # aggrs.append({row[RING_NODE]: anode})
                        aggrs[rid][row[RING_NODE]].update(anode)
                    else:
                        warning ("Node {} is already Aggregator for this (#{}) ring. Adding a new interface only".format(row[RING_NODE], rid))
                        print("aggrs for row {} & node {}:".format(rid, row[RING_NODE]))
                        print("aggrs for row {} & node {}: {}".format(rid, row[RING_NODE], aggrs[rid][row[RING_NODE]]))

                    aggrs[rid][row[RING_NODE]]['ports'].append({ 'port': row[RING_PORT], 'peer_name': row[RING_PEER], 'peer_port': row[RING_PEER_PORT], 'media': row[RING_PORT_TYPE], 'sfp': row[RING_SFP]})
                    aggrs[rid][row[RING_NODE]]['interfaces'].append({ 'name' : "{}_if1".format(row[RING_PEER]), 'port': row[RING_PORT], 'vlan': 1, 'ip' : "{}/31".format(row[RING_IP_ADDR]), 'ospf_area': int(row[RING_OSPF_AREA]), 'srlg': _srlg} )

                else:
                    warning("There is a node {} in the ring {} that is neither an Aggregator nor Distributed Switch! \n \
                    Resulted config might not be operational!!".format(row[RING_NODE], rid))

            else:
                # DS node
                print("Node: {}: ".format(row[RING_NODE]))
                print("@@ {} @@".format(dsnodes[rid][row[RING_NODE]]))
                dsnodes[rid][row[RING_NODE]]['ports'].append( {'port': row[RING_PORT], 'peer_name': row[RING_PEER], 'peer_port': row[RING_PEER_PORT], 'media': row[RING_PORT_TYPE], 'sfp': row[RING_SFP]})
                # Check SFP: QSFP28?
                if row[RING_SFP] == "QSFP28":
                    _tp = "{}/1".format(row[RING_PORT])
                    dsnodes[rid][row[RING_NODE]]['ports'].append( {'port': _tp, 'peer_name': row[RING_PEER], 'peer_port': row[RING_PEER_PORT], 'media': row[RING_PORT_TYPE], 'sfp': "CON"})
                    dsnodes[rid][row[RING_NODE]]['interfaces'].append( { 'name' : "{}_if1".format(row[RING_PEER]), 'port': _tp, 'vlan': 1, 'ip' : "{}/31".format(row[RING_IP_ADDR]), 'ospf_area': int(row[RING_OSPF_AREA]), 'srlg': _srlg} )
                else:
                    _tp = row[RING_PORT]
                    dsnodes[rid][row[RING_NODE]]['interfaces'].append( { 'name' : "{}_if1".format(row[RING_PEER]), 'port': row[RING_PORT], 'vlan': 1, 'ip' : "{}/31".format(row[RING_IP_ADDR]), 'ospf_area': int(row[RING_OSPF_AREA]), 'srlg': _srlg} )

                # Update site info
##            #>>>if peer_port in r[4])][r[4]]['site']
                print("#" + "-" * 50)
                print(dsnodes[rid][row[RING_NODE]])
                print("#" + "-" * 50)
                tmp_node = dsnodes[rid][row[RING_NODE]]
                if 'site' in tmp_node.keys():
                    print ("\tsite")
                    print ("\tds pair id: {}".format(tmp_node['site']['pair_id']))
                    print("\t{}".format(dspairs.keys()))
                    if int(tmp_node['site']['pair_id']) in dspairs.keys():
                        print("!\t\t{}".format(tmp_node['site']['pair_id']))
                        if row[RING_NODE] in dspairs[int(tmp_node['site']['pair_id'])]['nodes']:
                            print ("->\tNode {} with pair id {}".format(row[RING_NODE], tmp_node['site']['pair_id']))
                            if row[RING_PEER] in  dspairs[int(tmp_node['site']['pair_id'])]['nodes']:
                                tmp_node['site'].update({'site_id': row[RING_ID], 'peer_port': _tp, 'node_idx': row[RING_NODE_NUM]})

        current_row += 1
##        #print ("Incr: {}".format(ri))
##

def proc_aggrs_list(_rings_list):
    al = []
    with open(_rings_list, "r") as f:
        for i, line in enumerate(f):
            if i == 0:
                continue
#            print("i: {}; Aggr LINE: {}".format(i, line))
#            r = rw.search(line)
#            if not r:
#                #z = re.split(';', remove_control_chars(line.rstrip()), 10)
            line = re.split(';', str_norm(line.rstrip().lstrip()), 10)
            al.append (line)
    f.close()
    return al

def get_chassis(_type):
    sp = re.compile(r'7750')
    r = sp.search(_type)
    if r:
        return 7750
    sp = re.compile(r'7250')
    r = sp.search(_type)
    if r:
        return 7250

def get_timos_version(_timos):
    debug("Timos str in: {}".format(_timos), dbg)
    vs = '0.0.0'; v = {}
    v['full'] = ''; v['major'] = 0;  v['minor'] = 0;  v['release'] = 0;
    sp = re.compile(r'[tT][iI][mM][oO][sS](?=\-[a-zA-Z]+\-)(\d+)\.(\d+)\.(.*)|(?!\-[a-zA-Z]+\-)(\d+)\.(\d+)\.(.*)')
    r = sp.search(_timos.strip('\ '))
    if r and r.group(0):
        vs = r.group(0)
    v['full'] = vs.strip('\ ')
    v['major'] = vs.split('.')[0]
    va = vs.split('.')
    if len(va) > 2:
        v['minor'] = va[1]
        v['release'] = va[2]
    return v

#def save_yaml(_routers, out_yaml = None, marker = None, whole = 0):
def save_yaml(_routers, out_yaml = None, whole = 0):
    if out_yaml is None:
        out = sys.stdout
    else:
        if whole:
            out = open(out_yaml, 'a')
        else:
            out = open(out_yaml, 'w')
    try:
#        if marker:
#            out.write("#"*52)
#            out.write("# {:^52} #".format(marker))
#            out.write("#"*52)
        d = yaml.dump(_routers, out, indent=4, default_flow_style=False)
    finally:
        if out_yaml is not None:
            out.close()

if __name__ == '__main__':
    o = redopt()

    if not os.path.exists(os.path.abspath(o['aggrs'])):
        error("Coudn't find aggrs file \"{}\" !".format(o['aggrs']))
        sys.exit(2)

    if not os.path.exists(os.path.abspath(o['rings'])):
        error("Coudn't find rings file \"{}\" !".format(o['rings']))
        sys.exit(2)

##    # Aggr node List
##    info("Getting aggregation nodes data...")
##    al = proc_aggrs_list(o['aggrs'])
##
##    nodes_in_a_ring = []
##    aggrs_in_a_ring = []
##
##    # aggregators list
##    aggrs = {}
##    rings = {}
##    # Aggr routers structured data
##    daggr = {}
##    a_header_map = [ 'id', 'ip', 'name', 'sid', 'sbfd', 'chassis', 'sros', 'sdpid' ]
##    # Convert aggr_nodes list into dict
##    for a in al:
##    ###    0      1               2        3      4         5            6
##    ### ['1', '172.16.252.1', 'sr1-kie2', '1', '524288', '7750-SR12', 'TiMOS-C-19.10.R6', 1001]
##    ###   id      system_ip     name      sid   bfd        chassis      SROS              sdpid
##    #
##        if not (a[2]):
##            continue
##        data =  (dict(zip(a_header_map,a)))
##        daggr[a[2]] = data
##        debug ("AGGR data: {}".format(data), dbg)
##    al = ''
##

    daggr = {}
    aggr_nodes = {}
    wb = load_workbook(filename = o['aggrs'], read_only=True, data_only=True)

    #aggr_nodes = get_aggr_nodes_data (wb):
    daggr = get_aggr_nodes_data(wb)
    wb.close()

    dspairs = {}
    dsnodes = {}
    wb = load_workbook(filename = o['rings'], read_only=True, data_only=True)

    # DS Node List
    #if o['index']:
    dsnodes, dspairs = proc_dsnodes_list2(wb, o['index'] if 'index' in o else 0)
    #else:
    #    dsnodes, dspairs  = proc_dsnodes_list2(wb)
    print("#" + "-" * 50)
    print("DSNODES: {}:".format(dsnodes))
    print("DSPAIRS: {}:".format(dspairs))
    print("#" + "-" * 50)

    rings = {}
    aggrs = {}
    aggrs_in_a_ring = []
    proc_rings_list2(wb, o['index'] if 'index' in o else 0)

    print("#" + "-" * 50)
    print("DSNODES: {}:".format(dsnodes))
    print("DSPAIRS: {}:".format(dspairs))
    print("#" + "-" * 50)

## DEBUG
    print("aggrs: {}".format(aggrs))
    print("rings: {}".format(rings))
## DEBUG
    yaml_conf = {}
    for d in dsnodes.keys():
        if d not in yaml_conf.keys():
            yaml_conf[d] = {}
            yaml_conf[d]['routers'] = {}
        yaml_conf[d]['routers'].update({'ds': dsnodes[d]})
    for a in aggrs.keys():
        if a in yaml_conf.keys():
            yaml_conf[a]['routers'].update({'aggr': aggrs[a]})
        else:
            warning("Ring in aggragation data (#{}) doesn't match any in the DS one!".format(a))

    print("#" + "-" * 50)
    for y in yaml_conf.keys():
        if 'yaml' in o:
            print("Ring {} config".format(y))
            print (yaml_conf[y])
            save_yaml(yaml_conf[y], "{}-ring{}.yaml".format(o['yaml'], y))
        else:
            save_yaml(yaml_conf[y])
        print("#" + "-" * 50)

#vim: tabstop=8 expandtab shiftwidth=4 softtabstop=4 syntax=python
